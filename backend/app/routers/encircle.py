"""
Encircle XLSX import router.

Handles importing items and images from Encircle detailed XLSX export files.
Supports:
- Parent location extraction from merged cell E1-G3
- Sub-location (room) parsing from rows
- Additional item fields: Brand, Model, Serial, Quantity, Retailer, 
  Purchase date, Purchase price, Estimated value, Warranty info
- Hierarchical location structure (Parent → Sub-locations)
"""

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.orm import Session
from typing import List, Optional
from pathlib import Path
from uuid import UUID
import shutil
import re
import logging
import tempfile
from datetime import datetime, date, timezone
from io import BytesIO
from openpyxl import load_workbook
import base64

from .. import models, schemas
from ..deps import get_db
from ..config import settings

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/import", tags=["import"])

# Upload directory for photos
# Media files are stored in /app/data/media to ensure they persist with the database
UPLOAD_DIR = Path("/app/data/media/photos")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

# Allowed image extensions
ALLOWED_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".webp"}

# Column name patterns to identify header row
COL_NO = "No."
COL_NAME = "Description"

# Encircle parent location header position (columns D-G, rows 1-3)
# These are 1-indexed column numbers for Excel compatibility
PARENT_LOCATION_COL_START = 4  # Column D
PARENT_LOCATION_COL_END = 7    # Column G
PARENT_LOCATION_ROW_START = 1
PARENT_LOCATION_ROW_END = 3

# Precompiled regex for numeric prefix matching
NO_PREFIX_RE = re.compile(r"^0*(\d+)_", re.IGNORECASE)

# MIME type mapping for image extensions
MIME_TYPES = {
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".gif": "image/gif",
    ".webp": "image/webp",
}


def get_mime_type(path: Path) -> str:
    """Get MIME type for an image file based on its extension."""
    ext = path.suffix.lower()
    return MIME_TYPES.get(ext, "image/jpeg")


def normalize_text(s: str) -> str:
    """
    Normalize text for fuzzy matching:
        - lowercase
        - remove all non-alphanumeric chars
    """
    return re.sub(r"[^a-z0-9]+", "", s.lower())


def find_header_row(ws) -> int:
    """
    Find row index with actual headers (contains 'No.' and 'Description').
    """
    for row_idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True), start=1):
        vals = [
            str(v).strip().lower()
            for v in row
            if v is not None and str(v).strip() != ""
        ]
        if not vals:
            continue

        has_no = any(v in ("no.", "no", "#") for v in vals)
        has_desc = "description" in vals

        if has_no and has_desc:
            return row_idx

    raise ValueError(
        "Could not find header row containing 'No.' and 'Description'. "
        "Please confirm the Encircle export format."
    )


def get_column_indices(header_row: tuple) -> dict:
    """
    Map column names to their indices.
    Supports: No., Description, Brand, Model, Serial, QTY, Retailer,
    Purchase Date, Purchase Price, Estimated Value (RCV), Warranty Duration,
    Extended Warranty Policy, Warranty Phone
    """
    indices = {}
    for idx, value in enumerate(header_row):
        if value is None:
            continue
        val = str(value).strip().lower()
        if val in ("no.", "no", "#"):
            indices["no"] = idx
        elif val == "description":
            indices["description"] = idx
        elif val == "location":
            indices["location"] = idx
        elif val == "brand":
            indices["brand"] = idx
        elif val == "model":
            indices["model"] = idx
        elif val == "serial":
            indices["serial"] = idx
        elif val == "qty" or val == "quantity":
            indices["quantity"] = idx
        elif val == "retailer":
            indices["retailer"] = idx
        elif "purchase date" in val:
            indices["purchase_date"] = idx
        elif "purchase price" in val:
            indices["purchase_price"] = idx
        elif "rcv" in val or "replacement" in val or "estimated" in val:
            indices["estimated_value"] = idx
        elif "acv" in val or "actual" in val:
            indices["acv"] = idx
        elif "warranty duration" in val:
            indices["warranty_duration"] = idx
        elif "extended warranty" in val or "warranty policy" in val:
            indices["extended_warranty_policy"] = idx
        elif "warranty phone" in val:
            indices["warranty_phone"] = idx
    return indices


def extract_location_from_cell_value(cell_value: str) -> Optional[str]:
    """
    Extract location name from a cell value that may contain multiple lines.
    The Encircle format often has the location name on the first line,
    followed by "Report Date:" on subsequent lines.
    
    Returns the first valid line that doesn't look like metadata.
    """
    if not cell_value:
        return None
    
    # Split by newlines and check each line
    lines = str(cell_value).strip().split('\n')
    for line in lines:
        line = line.strip()
        if not line:
            continue
        # Skip lines that look like metadata
        line_lower = line.lower()
        if ("report date" in line_lower or 
            "date:" in line_lower or
            "encircle" in line_lower):
            continue
        # Found a valid location name
        return line
    
    return None


def extract_location_from_filename(filename: Optional[str]) -> Optional[str]:
    """
    Extract location name from an Encircle export filename.
    Encircle filenames typically follow the pattern: LocationName__Encircle_Detailed__12345.xlsx
    
    Returns the location name portion, or None if the pattern doesn't match.
    """
    if not filename:
        return None
    
    # Remove path if present, get just the filename
    name = Path(filename).stem
    
    # Pattern: LocationName__Encircle_Detailed__12345
    # Try to extract the part before "__Encircle"
    match = re.match(r"^(.+?)__Encircle", name, re.IGNORECASE)
    if match:
        location = match.group(1).strip()
        if location:
            # Replace single underscores with spaces for better readability
            # (double underscores are unlikely in location names before __Encircle)
            location = location.replace('_', ' ').strip()
            return location
    
    return None


def extract_parent_location_name(ws, filename: Optional[str] = None) -> Optional[str]:
    """
    Extract parent location name from merged cell E1-G3 or similar area.
    The Encircle format typically has the location name in the header area
    around cells E1-G3 (merged).
    
    Falls back to extracting location from filename if cell extraction fails.
    
    Args:
        ws: The worksheet to extract from
        filename: Optional filename to use as fallback for location extraction
    """
    # Check for merged cells in the parent location header range first
    for merged_range in ws.merged_cells.ranges:
        # Check if the merged range includes cells around the expected header area
        if (merged_range.min_row <= PARENT_LOCATION_ROW_END and 
            merged_range.max_row >= PARENT_LOCATION_ROW_START and
            merged_range.min_col >= PARENT_LOCATION_COL_START and 
            merged_range.min_col <= PARENT_LOCATION_COL_END):
            # Get the value from the top-left cell of the merged range
            cell_value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
            if cell_value:
                location_name = extract_location_from_cell_value(str(cell_value))
                if location_name:
                    return location_name
    
    # Fallback: scan cells in the parent location header range for a location name
    for row_idx in range(PARENT_LOCATION_ROW_START, PARENT_LOCATION_ROW_END + 1):
        for col_idx in range(PARENT_LOCATION_COL_START, PARENT_LOCATION_COL_END + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                location_name = extract_location_from_cell_value(str(cell_value))
                if location_name:
                    return location_name
    
    # Final fallback: try to extract location from filename
    if filename:
        location_name = extract_location_from_filename(filename)
        if location_name:
            return location_name
    
    return None


def is_room_header_row(row: tuple, no_idx: int, desc_idx: int) -> Optional[str]:
    """
    Check if a row is a room/sub-location header.
    Room headers typically have:
    - No value in the "No." column
    - A room name in column B (Description or before it)
    
    Returns the room name if it's a room header, None otherwise.
    """
    # Get the "No." column value
    no_value = row[no_idx] if no_idx < len(row) else None
    
    # If there's a number in the No. column, it's an item row, not a room header
    if no_value is not None and str(no_value).strip():
        try:
            int(str(no_value).strip().split(".")[0])
            return None  # Has a valid item number
        except (ValueError, TypeError):
            pass  # Not a valid number, continue checking
    
    # Check column B (index 1) for room name - this is typical for Encircle format
    # Room headers are in column B when there's no item number
    if len(row) > 1:
        col_b_value = row[1]  # Column B is index 1
        if col_b_value and str(col_b_value).strip():
            room_name = str(col_b_value).strip()
            # Exclude if it looks like an item description or header text
            if (len(room_name) > 0 and 
                room_name.lower() not in ("description", "no.", "no", "#", "photo", "brand", "model") and
                not room_name.startswith("Report Date")):
                return room_name
    
    return None


def parse_date(value) -> Optional[date]:
    """Parse a date value from various formats."""
    if value is None:
        return None
    
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    
    val_str = str(value).strip()
    if not val_str:
        return None
    
    # Try various date formats
    formats = [
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%b %d, %Y",
        "%B %d, %Y",
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(val_str, fmt).date()
        except ValueError:
            continue
    
    return None


def parse_currency(value) -> Optional[float]:
    """Parse a currency value, removing symbols and handling formatting."""
    if value is None:
        return None
    
    if isinstance(value, (int, float)):
        return float(value)
    
    val_str = str(value).strip()
    if not val_str:
        return None
    
    # Check if negative (handle parentheses or minus sign)
    is_negative = val_str.startswith('-') or val_str.startswith('(')
    
    # Remove currency symbols and commas, preserve digits, dots, and minus signs
    val_str = re.sub(r'[^\d.-]', '', val_str)
    
    # Handle multiple decimal points
    if val_str.count('.') > 1:
        parts = val_str.split('.')
        val_str = parts[0] + '.' + ''.join(parts[1:])
    
    # Remove any extra minus signs (keep only leading)
    if val_str.count('-') > 1:
        val_str = '-' + val_str.replace('-', '')
    
    if val_str and val_str != '.' and val_str != '-':
        try:
            result = float(val_str)
            # For inventory values, we typically don't want negative values
            # but we'll preserve them in case they're meaningful
            return result
        except ValueError:
            pass
    
    return None


def parse_warranty_duration(value) -> Optional[int]:
    """Parse warranty duration, returning months."""
    if value is None:
        return None
    
    val_str = str(value).strip().lower()
    if not val_str:
        return None
    
    # Try to extract number
    match = re.search(r'(\d+)', val_str)
    if not match:
        return None
    
    num = int(match.group(1))
    
    # Check for year indicators
    if 'year' in val_str:
        return num * 12
    
    # Default to months
    return num


def match_images_by_name(images: List[Path], description: str) -> List[Path]:
    """
    Match images whose normalized filename contains the normalized description.
    """
    desc_norm = normalize_text(description)
    if not desc_norm:
        return []

    matches: List[Path] = []
    for path in images:
        stem_norm = normalize_text(path.stem)
        if desc_norm in stem_norm:
            matches.append(path)
    return matches


def match_images_by_number(images: List[Path], item_no: int) -> List[Path]:
    """
    Match images by numeric prefix pattern (e.g., 0003_description.jpg).
    """
    matches: List[Path] = []
    for path in images:
        m = NO_PREFIX_RE.match(path.name)
        if m:
            try:
                no = int(m.group(1))
                if no == item_no:
                    matches.append(path)
            except ValueError:
                continue
    return matches


def classify_image_type(path: Path) -> str:
    """
    Map filename to photo type.
    """
    name = path.name.lower()
    if "receipt" in name:
        return "receipt"
    if "data_tag" in name or "datatag" in name or "tag" in name:
        return "data_tag"
    return "default"


def estimate_value_from_image(image_path: Path) -> tuple[Optional[float], Optional[str], Optional[str], Optional[str]]:
    """
    Use Gemini AI to analyze an image (data tag or item photo) and extract:
    - Estimated value
    - Model number
    - Serial number
    - Brand
    
    Returns (estimated_value, model_number, serial_number, brand) or (None, None, None, None) if AI is not configured or fails.
    """
    if not settings.GEMINI_API_KEY or not settings.GEMINI_API_KEY.strip():
        return None, None, None, None
    
    try:
        import google.generativeai as genai
        
        genai.configure(api_key=settings.GEMINI_API_KEY)
        
        # Read and encode the image
        with open(image_path, "rb") as f:
            image_data = f.read()
        image_base64 = base64.b64encode(image_data).decode("utf-8")
        
        model = genai.GenerativeModel(settings.GEMINI_MODEL)
        
        # Determine MIME type
        mime_type = get_mime_type(image_path)
        
        prompt = """Analyze this image which may be a product data tag, label, or photo of an item.

Extract the following information if visible or identifiable:
1. brand: The brand or manufacturer name
2. model_number: The model number or part number  
3. serial_number: The serial number (S/N)
4. estimated_value: Based on the brand, model, and product type, estimate the current market/replacement value in USD (just the number)

Return ONLY a JSON object with these fields. Use null for any field that cannot be determined.

Example format:
{
  "brand": "GE",
  "model_number": "GIE19JSNBRSS",
  "serial_number": "SR769052",
  "estimated_value": 850
}"""

        image_part = {
            "mime_type": mime_type,
            "data": image_base64
        }
        
        response = model.generate_content([prompt, image_part])
        response_text = response.text
        
        # Parse the response
        import json
        json_match = re.search(r'\{[\s\S]*\}', response_text)
        if json_match:
            parsed = json.loads(json_match.group())
            
            estimated_value = None
            value_str = parsed.get("estimated_value") or parsed.get("value")
            if value_str:
                try:
                    if isinstance(value_str, (int, float)):
                        estimated_value = float(value_str)
                    else:
                        clean_value = re.sub(r'[^\d.]', '', str(value_str))
                        if clean_value:
                            estimated_value = float(clean_value)
                except (ValueError, TypeError):
                    pass
            
            model_number = parsed.get("model_number") or parsed.get("model")
            serial_number = parsed.get("serial_number") or parsed.get("serial")
            brand = parsed.get("brand") or parsed.get("manufacturer")
            
            return estimated_value, model_number, serial_number, brand
            
    except Exception as e:
        logger.warning(f"Failed to estimate value from image: {e}")
    
    return None, None, None, None


def estimate_value_from_description(
    name: str,
    brand: Optional[str] = None,
    model_number: Optional[str] = None,
    description: Optional[str] = None
) -> Optional[float]:
    """
    Use Gemini AI to estimate the value of an item based on its name, brand, model, and description.
    
    Returns the estimated value or None if AI is not configured or fails.
    """
    if not settings.GEMINI_API_KEY or not settings.GEMINI_API_KEY.strip():
        return None
    
    try:
        import google.generativeai as genai
        
        genai.configure(api_key=settings.GEMINI_API_KEY)
        model = genai.GenerativeModel(settings.GEMINI_MODEL)
        
        # Build context about the item
        item_info = f"Item: {name}"
        if brand:
            item_info += f"\nBrand: {brand}"
        if model_number:
            item_info += f"\nModel: {model_number}"
        if description:
            item_info += f"\nDescription: {description}"
        
        prompt = f"""Based on the following item information, estimate the current market/replacement value in USD.

{item_info}

Return ONLY a JSON object with the estimated value. Example:
{{"estimated_value": 450}}

If you cannot make a reasonable estimate, return:
{{"estimated_value": null}}"""

        response = model.generate_content(prompt)
        response_text = response.text
        
        # Parse the response
        import json
        json_match = re.search(r'\{[\s\S]*\}', response_text)
        if json_match:
            parsed = json.loads(json_match.group())
            
            value_str = parsed.get("estimated_value") or parsed.get("value")
            if value_str:
                try:
                    if isinstance(value_str, (int, float)):
                        return float(value_str)
                    else:
                        clean_value = re.sub(r'[^\d.]', '', str(value_str))
                        if clean_value:
                            return float(clean_value)
                except (ValueError, TypeError):
                    pass
                    
    except Exception as e:
        logger.warning(f"Failed to estimate value from description: {e}")
    
    return None


class ImportResult:
    def __init__(self):
        self.items_created = 0
        self.photos_attached = 0
        self.items_without_photos = 0
        self.locations_created = 0
        self.sublocations_created = 0
        self.parent_location_name: Optional[str] = None
        self.errors: List[str] = []
        self.log: List[str] = []


def process_encircle_import(
    xlsx_content: bytes,
    images: List[tuple],  # List of (filename, content) tuples
    db: Session,
    match_by_name: bool = True,
    parent_location_id: Optional[str] = None,
    create_parent_from_file: bool = True,
    xlsx_filename: Optional[str] = None
) -> ImportResult:
    """
    Process Encircle XLSX import with optional images.
    
    Args:
        xlsx_content: Bytes content of the XLSX file
        images: List of (filename, content) tuples for images
        db: Database session
        match_by_name: If True, match images by Description; else by No. prefix
        parent_location_id: Optional ID of existing parent location to use
        create_parent_from_file: If True and no parent_location_id, create parent from file header
        xlsx_filename: Optional filename for fallback location extraction
    
    Returns:
        ImportResult with details about the import
    """
    result = ImportResult()
    
    try:
        # Load workbook from bytes
        wb = load_workbook(filename=BytesIO(xlsx_content), data_only=True)
        ws = wb.active
        
        # Extract parent location name from the file (merged cell E1-G3, with filename fallback)
        file_parent_location_name = extract_parent_location_name(ws, xlsx_filename)
        result.parent_location_name = file_parent_location_name
        if file_parent_location_name:
            result.log.append(f"Detected parent location in file: {file_parent_location_name}")
        
        # Find header row
        header_row_idx = find_header_row(ws)
        result.log.append(f"Detected header row at row {header_row_idx}")
        
        # Get header row and column indices
        header_row = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))[0]
        col_indices = get_column_indices(header_row)
        
        if "no" not in col_indices or "description" not in col_indices:
            raise ValueError("Required columns 'No.' and 'Description' not found in header row")
        
        result.log.append(f"Column mapping: {col_indices}")
        
        # Create temp directory for images if needed
        image_paths: List[Path] = []
        temp_dir = None
        
        if images:
            temp_dir = Path(tempfile.mkdtemp())
            for filename, content in images:
                img_path = temp_dir / filename
                with open(img_path, "wb") as f:
                    f.write(content)
                image_paths.append(img_path)
            result.log.append(f"Indexed {len(image_paths)} images for matching")
        
        # Determine parent location
        parent_location: Optional[models.Location] = None
        parent_location_db_id = None
        
        if parent_location_id:
            # Use the provided parent location
            try:
                parent_uuid = UUID(parent_location_id)
                parent_location = db.query(models.Location).filter(
                    models.Location.id == parent_uuid
                ).first()
                if parent_location:
                    parent_location_db_id = parent_location.id
                    result.log.append(f"Using existing parent location: {parent_location.name}")
                else:
                    result.log.append(f"Warning: Provided parent location ID not found, will create from file")
            except (ValueError, TypeError):
                result.log.append(f"Warning: Invalid parent location ID format, will create from file")
        
        if parent_location_db_id is None and create_parent_from_file and file_parent_location_name:
            # Create parent location from file header
            existing_parent = db.query(models.Location).filter(
                models.Location.name == file_parent_location_name,
                models.Location.parent_id == None
            ).first()
            
            if existing_parent:
                parent_location = existing_parent
                parent_location_db_id = existing_parent.id
                result.log.append(f"Using existing parent location from file: {file_parent_location_name}")
            else:
                parent_location = models.Location(
                    name=file_parent_location_name,
                    is_primary_location=True
                )
                db.add(parent_location)
                db.flush()
                parent_location_db_id = parent_location.id
                result.locations_created += 1
                result.log.append(f"Created parent location from file: {file_parent_location_name}")
        
        # Cache for sub-locations (rooms)
        sublocations_cache: dict = {}
        
        # Load existing sub-locations for this parent into cache
        if parent_location_db_id:
            existing_sublocations = db.query(models.Location).filter(
                models.Location.parent_id == parent_location_db_id
            ).all()
            for loc in existing_sublocations:
                sublocations_cache[loc.name.lower()] = loc.id
        
        # Track current sub-location (room) while processing rows
        current_sublocation_id = parent_location_db_id
        current_sublocation_name = file_parent_location_name or "Root"
        
        # Process data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
            # Skip empty rows
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            
            # Check if this is a room/sub-location header row
            room_name = is_room_header_row(row, col_indices["no"], col_indices["description"])
            if room_name:
                # This is a room header - update current sub-location
                room_key = room_name.lower()
                if room_key not in sublocations_cache:
                    # Create new sub-location
                    new_sublocation = models.Location(
                        name=room_name,
                        parent_id=parent_location_db_id
                    )
                    db.add(new_sublocation)
                    db.flush()
                    sublocations_cache[room_key] = new_sublocation.id
                    result.sublocations_created += 1
                    result.log.append(f"Created sub-location (room): {room_name}")
                
                current_sublocation_id = sublocations_cache[room_key]
                current_sublocation_name = room_name
                continue
            
            # Extract item data
            raw_no = row[col_indices["no"]] if "no" in col_indices else None
            raw_name = row[col_indices["description"]] if "description" in col_indices else None
            
            if raw_no is None or raw_name is None:
                continue
            
            # Parse item number
            try:
                no = int(str(raw_no).strip().split(".")[0])
            except (TypeError, ValueError):
                continue
            
            name = str(raw_name).strip()
            if not name:
                continue
            
            # Extract additional fields
            brand = None
            if "brand" in col_indices and col_indices["brand"] < len(row):
                val = row[col_indices["brand"]]
                if val:
                    brand = str(val).strip() or None
            
            model_number = None
            if "model" in col_indices and col_indices["model"] < len(row):
                val = row[col_indices["model"]]
                if val:
                    model_number = str(val).strip() or None
            
            serial_number = None
            if "serial" in col_indices and col_indices["serial"] < len(row):
                val = row[col_indices["serial"]]
                if val:
                    serial_number = str(val).strip() or None
            
            retailer = None
            if "retailer" in col_indices and col_indices["retailer"] < len(row):
                val = row[col_indices["retailer"]]
                if val:
                    retailer = str(val).strip() or None
            
            purchase_date = None
            if "purchase_date" in col_indices and col_indices["purchase_date"] < len(row):
                purchase_date = parse_date(row[col_indices["purchase_date"]])
            
            purchase_price = None
            if "purchase_price" in col_indices and col_indices["purchase_price"] < len(row):
                purchase_price = parse_currency(row[col_indices["purchase_price"]])
            
            estimated_value = None
            estimated_value_ai_date = None
            if "estimated_value" in col_indices and col_indices["estimated_value"] < len(row):
                estimated_value = parse_currency(row[col_indices["estimated_value"]])
            
            # AI estimation: If estimated_value is None/blank, try to get it from AI
            if estimated_value is None and image_paths:
                # First, find data tag images for this item
                if match_by_name:
                    matched_images = match_images_by_name(image_paths, name)
                    if not matched_images:
                        matched_images = match_images_by_number(image_paths, no)
                else:
                    matched_images = match_images_by_number(image_paths, no)
                
                # Look for data tag photos first
                data_tag_photos = [img for img in matched_images if classify_image_type(img) == "data_tag"]
                
                if data_tag_photos:
                    # Try to get value from data tag photo
                    for dt_photo in data_tag_photos:
                        ai_value, ai_model, ai_serial, ai_brand = estimate_value_from_image(dt_photo)
                        if ai_value is not None:
                            estimated_value = ai_value
                            estimated_value_ai_date = datetime.now(timezone.utc).strftime("%m/%d/%y")
                            result.log.append(f"  -> AI estimated value ${ai_value:.2f} from data tag photo")
                            # Also update model/serial/brand if they were blank
                            if not model_number and ai_model:
                                model_number = ai_model
                            if not serial_number and ai_serial:
                                serial_number = ai_serial
                            if not brand and ai_brand:
                                brand = ai_brand
                            break
                
                # If still no value, try any other matched photos
                if estimated_value is None and matched_images:
                    other_photos = [img for img in matched_images if classify_image_type(img) != "data_tag"]
                    for photo in other_photos:
                        ai_value, ai_model, ai_serial, ai_brand = estimate_value_from_image(photo)
                        if ai_value is not None:
                            estimated_value = ai_value
                            estimated_value_ai_date = datetime.now(timezone.utc).strftime("%m/%d/%y")
                            result.log.append(f"  -> AI estimated value ${ai_value:.2f} from item photo")
                            # Also update model/serial/brand if they were blank
                            if not model_number and ai_model:
                                model_number = ai_model
                            if not serial_number and ai_serial:
                                serial_number = ai_serial
                            if not brand and ai_brand:
                                brand = ai_brand
                            break
            
            # If still no estimated value, try AI estimation from item details (best guess)
            if estimated_value is None:
                ai_value = estimate_value_from_description(name, brand, model_number, None)
                if ai_value is not None:
                    estimated_value = ai_value
                    estimated_value_ai_date = datetime.now(timezone.utc).strftime("%m/%d/%y")
                    result.log.append(f"  -> AI estimated value ${ai_value:.2f} from item description")
            
            # Build warranties list if warranty info is present
            warranties = None
            warranty_duration = None
            extended_warranty_policy = None
            warranty_phone = None
            
            if "warranty_duration" in col_indices and col_indices["warranty_duration"] < len(row):
                warranty_duration = parse_warranty_duration(row[col_indices["warranty_duration"]])
            
            if "extended_warranty_policy" in col_indices and col_indices["extended_warranty_policy"] < len(row):
                val = row[col_indices["extended_warranty_policy"]]
                if val:
                    extended_warranty_policy = str(val).strip() or None
            
            if "warranty_phone" in col_indices and col_indices["warranty_phone"] < len(row):
                val = row[col_indices["warranty_phone"]]
                if val:
                    warranty_phone = str(val).strip() or None
            
            if warranty_duration or extended_warranty_policy or warranty_phone:
                warranties = []
                if warranty_duration:
                    warranties.append({
                        "type": "manufacturer",
                        "duration_months": warranty_duration,
                        "notes": f"Warranty phone: {warranty_phone}" if warranty_phone else None
                    })
                if extended_warranty_policy:
                    warranties.append({
                        "type": "extended",
                        "policy_number": extended_warranty_policy,
                        "notes": f"Warranty phone: {warranty_phone}" if warranty_phone else None
                    })
            
            # Create item with current sub-location
            item = models.Item(
                name=name,
                location_id=current_sublocation_id,
                brand=brand,
                model_number=model_number,
                serial_number=serial_number,
                retailer=retailer,
                purchase_date=purchase_date,
                purchase_price=purchase_price,
                estimated_value=estimated_value,
                estimated_value_ai_date=estimated_value_ai_date,
                warranties=warranties
            )
            db.add(item)
            db.flush()
            result.items_created += 1
            
            # Match and attach images
            if image_paths:
                if match_by_name:
                    matched_images = match_images_by_name(image_paths, name)
                    if not matched_images:
                        # Fallback to number matching
                        matched_images = match_images_by_number(image_paths, no)
                else:
                    matched_images = match_images_by_number(image_paths, no)
                
                if matched_images:
                    is_first = True
                    for img_path in matched_images:
                        # Generate unique filename
                        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S_%f")
                        file_extension = img_path.suffix or ".jpg"
                        new_filename = f"{item.id}_{timestamp}{file_extension}"
                        dest_path = UPLOAD_DIR / new_filename
                        
                        # Copy image to uploads
                        shutil.copy2(img_path, dest_path)
                        
                        # Determine photo type and MIME type
                        photo_type = classify_image_type(img_path)
                        mime_type = get_mime_type(img_path)
                        is_primary = is_first and photo_type == "default"
                        is_data_tag = photo_type == "data_tag"
                        
                        # Create photo record
                        photo = models.Photo(
                            item_id=item.id,
                            path=f"/uploads/photos/{new_filename}",
                            mime_type=mime_type,
                            is_primary=is_primary,
                            is_data_tag=is_data_tag,
                            photo_type=photo_type
                        )
                        db.add(photo)
                        result.photos_attached += 1
                        is_first = False
                    
                    result.log.append(f"Item #{no}: {name} ({current_sublocation_name}) -> {len(matched_images)} photo(s)")
                else:
                    result.items_without_photos += 1
                    result.log.append(f"Item #{no}: {name} ({current_sublocation_name}) -> no photos matched")
            else:
                result.items_without_photos += 1
        
        db.commit()
        
        # Cleanup temp directory
        if temp_dir and temp_dir.exists():
            try:
                shutil.rmtree(temp_dir)
            except Exception as cleanup_err:
                logger.warning(f"Failed to cleanup temp directory {temp_dir}: {cleanup_err}")
        
        result.log.append("--- Import Summary ---")
        result.log.append(f"Items created: {result.items_created}")
        result.log.append(f"Photos attached: {result.photos_attached}")
        result.log.append(f"Items without photos: {result.items_without_photos}")
        result.log.append(f"Parent locations created: {result.locations_created}")
        result.log.append(f"Sub-locations (rooms) created: {result.sublocations_created}")
        
    except Exception as e:
        db.rollback()
        result.errors.append("An internal error occurred during import.")
        result.log.append("Import failed due to an internal error.")
        logger.exception("Encircle import failed")
    
    return result


def preview_encircle_import(xlsx_content: bytes, filename: Optional[str] = None) -> dict:
    """
    Preview an Encircle XLSX file to extract the parent location name
    without actually importing anything.
    
    Args:
        xlsx_content: Bytes content of the XLSX file
        filename: Optional filename for fallback location extraction
    
    Returns:
        Dictionary with parent_location_name
    """
    try:
        wb = load_workbook(filename=BytesIO(xlsx_content), data_only=True)
        ws = wb.active
        parent_location_name = extract_parent_location_name(ws, filename)
        return {
            "parent_location_name": parent_location_name,
            "success": True
        }
    except Exception as e:
        logger.exception("Failed to preview Encircle file")
        return {
            "parent_location_name": None,
            "success": False,
            "error": "An error occurred while previewing the Encircle file."
        }


@router.post("/encircle/preview")
async def preview_encircle(
    xlsx_file: UploadFile = File(..., description="Encircle XLSX export file to preview")
):
    """
    Preview an Encircle XLSX file to extract the parent location name
    without importing anything. Use this to show users what location
    will be created before they confirm the import.
    """
    # Validate file type
    if not xlsx_file.filename or not xlsx_file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="Invalid file type. Please upload an XLSX file."
        )
    
    # Read xlsx content
    xlsx_content = await xlsx_file.read()
    
    # Preview the file (pass filename for fallback location extraction)
    result = preview_encircle_import(xlsx_content, xlsx_file.filename)
    
    if not result["success"]:
        raise HTTPException(
            status_code=400,
            detail=f"Failed to preview file: {result.get('error', 'Unknown error')}"
        )
    
    return {
        "parent_location_name": result["parent_location_name"]
    }


@router.post("/encircle")
async def import_encircle(
    xlsx_file: UploadFile = File(..., description="Encircle XLSX export file"),
    images: Optional[List[UploadFile]] = File(None, description="Image files to import"),
    match_by_name: bool = Form(True, description="Match images by description name"),
    parent_location_id: Optional[str] = Form(None, description="ID of existing parent location to use"),
    create_parent_from_file: bool = Form(True, description="Create parent location from file if no parent_location_id provided"),
    db: Session = Depends(get_db)
):
    """
    Import items from an Encircle detailed XLSX export file.
    
    Features:
    - Extracts parent location name from merged cell E1-G3 (e.g., "Maine Cottage")
    - Parses sub-locations (rooms) from rows in column B
    - Imports additional fields: Brand, Model, Serial, Retailer, Purchase date,
      Purchase price, Estimated value, Warranty duration, Extended warranty policy,
      Warranty phone
    - Creates hierarchical location structure (Parent → Sub-locations/Rooms)
    
    Options:
    - parent_location_id: Use an existing location as parent instead of creating new
    - create_parent_from_file: If True and no parent_location_id, create parent from file header
    - match_by_name: Match images by description name (default) or by numeric prefix
    
    Returns import statistics and log.
    """
    # Validate file type
    if not xlsx_file.filename or not xlsx_file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="Invalid file type. Please upload an XLSX file."
        )
    
    # Read xlsx content
    xlsx_content = await xlsx_file.read()
    
    # Read images if provided
    image_data: List[tuple] = []
    if images:
        for img in images:
            if img.filename:
                ext = Path(img.filename).suffix.lower()
                if ext in ALLOWED_IMAGE_EXTENSIONS:
                    content = await img.read()
                    image_data.append((img.filename, content))
    
    # Process import (pass filename for fallback location extraction)
    result = process_encircle_import(
        xlsx_content=xlsx_content,
        images=image_data,
        db=db,
        match_by_name=match_by_name,
        parent_location_id=parent_location_id,
        create_parent_from_file=create_parent_from_file,
        xlsx_filename=xlsx_file.filename
    )
    
    if result.errors:
        raise HTTPException(
            status_code=400,
            detail={
                "message": "Import failed",
                "errors": result.errors,
                "log": result.log
            }
        )
    
    return {
        "message": "Import completed successfully",
        "items_created": result.items_created,
        "photos_attached": result.photos_attached,
        "items_without_photos": result.items_without_photos,
        "locations_created": result.locations_created,
        "sublocations_created": result.sublocations_created,
        "parent_location_name": result.parent_location_name,
        "log": result.log
    }
